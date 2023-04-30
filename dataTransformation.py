import pandas as pd
import re
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

NUM_ROUNDS = 10

# read in the original data
df = pd.read_excel('ExperimentCSVFiles/_experiment.xls')

# split ParticipantId and keep only the truncated version
df['ParticipantId'] = df['ParticipantId'].astype(str).str.slice(start=8, stop=-2).astype(int)

# split SubjectId and OpponentId to get SubjectRace, SubjectGender, OpponentRace, OpponentGender
df[['SubjectRace', 'SubjectGender']] = df['SubjectId'].str.split('_', expand=True)
df[['OpponentRace', 'OpponentGender']] = df['OpponentId'].str.split('_', expand=True)


# create a new DataFrame to hold the transformed data
new_df = pd.DataFrame(columns=[
    'Participant ID', 
    'Phase Number', 
    'Subject Self Identified Race', 
    'Subject Self Identified Gender',
    'Subject Assigned Race',
    'Subject Assigned Gender',
    'Opponent Assigned Race', 
    'Opponent Assigned Gender', 
    'Subject Round Score', 
    'Opponent Round Score',
    'Subject Total Score', 
    'Opponent Total Score', 
    'Cooperation Percentage',
])

# loop through each participant and phase
# group the rows by ParticipantId and Phase
grouped = df.groupby(['ParticipantId', 'Phase'])

# loop through the groups and perform the calculations
for (pid, ph), subset in grouped:
    
    # calculate the cumulative total scores
    subject_score = subset.loc[subset['RoundNumber']==NUM_ROUNDS, 'SubjectTotalScore'].iloc[0]
    opponent_score = subset.loc[subset['RoundNumber']==NUM_ROUNDS, 'OpponentTotalScore'].iloc[0]
        
    # count the number of times the subject cooperated
    cooperation_count = (subset['SubjectChoice'] == 0).sum()
        
    # calculate the percentage of times the subject cooperated
    cooperation_percentage = cooperation_count / NUM_ROUNDS
        
    # get the subject's self-race and self-gender
    subject_self_race = subset['SubjectSelfRace'].iloc[0]
    subject_self_gender = subset['SubjectSelfGender'].iloc[0]

    opponent_race = subset['OpponentRace'].iloc[0]
    opponent_gender = subset['OpponentGender'].iloc[0]

    subject_race = subset['SubjectRace'].iloc[0]
    subject_gender = subset['SubjectGender'].iloc[0]
        
    # get the subject's and opponent's round scores for round 10
    subject_round_score = subset.loc[subset['RoundNumber']==NUM_ROUNDS, 'SubjectRoundScore'].iloc[0]
    opponent_round_score = subset.loc[subset['RoundNumber']==NUM_ROUNDS, 'OpponentRoundScore'].iloc[0]

    # add a new row to the transformed DataFrame
    new_row = pd.DataFrame({
        'Participant ID': [pid],
        'Phase Number': [ph],
        'Subject Self Identified Race': [subject_self_race],
        'Subject Self Identified Gender': [subject_self_gender],
        'Subject Assigned Race': [subject_race],
        'Subject Assigned Gender': [subject_gender],
        'Opponent Assigned Race': [opponent_race],
        'Opponent Assigned Gender': [opponent_gender],
        'Subject Round Score': [subject_round_score],
        'Opponent Round Score': [opponent_round_score],
        'Subject Total Score': [subject_score],
        'Opponent Total Score': [opponent_score],
        'Cooperation Percentage': [cooperation_percentage],
    })
    new_df = pd.merge(new_df, new_row, how='outer')

new_df.to_excel('transformed_data.xlsx', index=False)

workbook = openpyxl.Workbook()
worksheet = workbook.active

for r in dataframe_to_rows(new_df, index=False, header=True):
    worksheet.append(r)

def convert_to_argb(hex_color):
    hex_color = re.sub(r'#(\w{2})(\w{2})(\w{2})', r'\3\2\1', hex_color)
    hex_color = 'FF' + hex_color
    return hex_color.upper()

color_list = ['#F2F2F2', '#D9D9D9']
color_list = ['FF' + re.sub(r'#(\w{2})(\w{2})(\w{2})', r'\3\2\1', c).upper() for c in color_list]

participant_rows = {}
for row in worksheet.iter_rows(min_row=2):
    participant_id = row[1].value
    participant_rows.setdefault(participant_id, []).append(row)


for rows in participant_rows.values():
    color_index = 0
    for i, row in enumerate(rows):
        if i % 1 == 0:
            color_index = (color_index + 1) % len(color_list)
        fill = PatternFill(start_color=color_list[color_index], end_color=color_list[color_index], fill_type='solid')
        for cell in row:
            cell.fill = fill

for column in worksheet.columns:
    max_length = 0
    column_letter = column[0].column_letter
    for cell in column:
        if len(str(cell.value)) > max_length:
            max_length = len(str(cell.value))
    adjusted_width = (max_length + 2)
    worksheet.column_dimensions[column_letter].width = adjusted_width

workbook.save('transformed_data_styled.xlsx')


