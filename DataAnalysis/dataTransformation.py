import pandas as pd
import re
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
import statsmodels.formula.api as smf


NUM_ROUNDS = 10

# read in the original data
df = pd.read_excel('ExperimentCSVFiles/_experiment.xls')

# split ParticipantId and keep only the truncated version
df['ParticipantId'] = df['ParticipantId'].astype(str).str.slice(start=8, stop=-2).astype(int)

# split SubjectId and OpponentId to get SubjectRace, SubjectGender, OpponentRace, OpponentGender
df[['SubjectRace', 'SubjectGender']] = df['SubjectId'].str.split('_', expand=True)
df[['OpponentRace', 'OpponentGender']] = df['OpponentId'].str.split('_', expand=True)


# create a new DataFrame to hold the transformed data
new_df = pd.DataFrame(columns=[
    'ParticipantID', 
    'PhaseNumber', 
    'SubjectSelfRace', 
    'SubjectSelfGender',
    'SubjectAvatarRace',
    'SubjectAvatarGender',
    'OpponentAvatarRace', 
    'OpponentAvatarGender', 
    'SubjectRoundScore', 
    'OpponentRoundScore',
    'SubjectTotalScore', 
    'OpponentTotalScore', 
    'CooperationPercentage',
])

# loop through each participant and phase
# group the rows by ParticipantId and Phase
grouped = df.groupby(['ParticipantId', 'Phase'])

# loop through the groups and perform the calculations
for (pid, ph), subset in grouped:
    
    # calculate the cumulative total scores
    subject_score = subset.loc[subset['RoundNumber']==NUM_ROUNDS, 'SubjectTotalScore'].iloc[0]
    opponent_score = subset.loc[subset['RoundNumber']==NUM_ROUNDS, 'OpponentTotalScore'].iloc[0]
        
    # count the number of times the subject cooperated
    cooperation_count = (subset['SubjectChoice'] == 0).sum()
        
    # calculate the percentage of times the subject cooperated
    cooperation_percentage = cooperation_count / float(NUM_ROUNDS)
        
    # get the subject's self-race and self-gender
    subject_self_race = subset['SubjectSelfRace'].iloc[0]
    subject_self_gender = subset['SubjectSelfGender'].iloc[0]

    opponent_race = subset['OpponentRace'].iloc[0]
    opponent_gender = subset['OpponentGender'].iloc[0]

    subject_race = subset['SubjectRace'].iloc[0]
    subject_gender = subset['SubjectGender'].iloc[0]
        
    # get the subject's and opponent's round scores for round 10
    subject_round_score = subset.loc[subset['RoundNumber']==NUM_ROUNDS, 'SubjectRoundScore'].iloc[0]
    opponent_round_score = subset.loc[subset['RoundNumber']==NUM_ROUNDS, 'OpponentRoundScore'].iloc[0]

    # add a new row to the transformed DataFrame
    new_row = pd.DataFrame({
        'ParticipantID': [pid],
        'PhaseNumber': [ph],
        'SubjectSelfRace': [subject_self_race],
        'SubjectSelfGender': [subject_self_gender],
        'SubjectAvatarRace': [subject_race],
        'SubjectAvatarGender': [subject_gender],
        'OpponentAvatarRace': [opponent_race],
        'OpponentAvatarGender': [opponent_gender],
        'SubjectRoundScore': [subject_round_score],
        'OpponentRoundScore': [opponent_round_score],
        'SubjectTotalScore': [subject_score],
        'OpponentTotalScore': [opponent_score],
        'CooperationPercentage': [cooperation_percentage],
    })
    new_df = pd.merge(new_df, new_row, how='outer')

new_df.to_excel('transformed_data.xlsx', index=False)

workbook = openpyxl.Workbook()
worksheet = workbook.active

for r in dataframe_to_rows(new_df, index=False, header=True):
    worksheet.append(r)

def convert_to_argb(hex_color):
    hex_color = re.sub(r'#(\w{2})(\w{2})(\w{2})', r'\3\2\1', hex_color)
    hex_color = 'FF' + hex_color
    return hex_color.upper()

color_list = ['#F2F2F2', '#D9D9D9']
color_list = ['FF' + re.sub(r'#(\w{2})(\w{2})(\w{2})', r'\3\2\1', c).upper() for c in color_list]

participant_rows = {}
for row in worksheet.iter_rows(min_row=2):
    participant_id = row[1].value
    participant_rows.setdefault(participant_id, []).append(row)


for rows in participant_rows.values():
    color_index = 0
    for i, row in enumerate(rows):
        if i % 1 == 0:
            color_index = (color_index + 1) % len(color_list)
        fill = PatternFill(start_color=color_list[color_index], end_color=color_list[color_index], fill_type='solid')
        for cell in row:
            cell.fill = fill

for column in worksheet.columns:
    max_length = 0
    column_letter = column[0].column_letter
    for cell in column:
        cell.alignment = Alignment(horizontal='center')
        if len(str(cell.value)) > max_length:
            max_length = len(str(cell.value))
    adjusted_width = (max_length + 2)
    worksheet.column_dimensions[column_letter].width = adjusted_width

# fit a regression model
model = smf.ols(formula='CooperationPercentage ~ C(SubjectSelfRace == OpponentAvatarRace) + C(SubjectAvatarRace == OpponentAvatarRace) + C(SubjectSelfGender == OpponentAvatarGender) + C(SubjectAvatarGender == OpponentAvatarGender) + ParticipantID', data=new_df)
results = model.fit()

# save the results to a file
with open('results.txt', 'w') as f:
    f.write(results.summary().as_text())


workbook.save('transformed_data_styled.xlsx')


